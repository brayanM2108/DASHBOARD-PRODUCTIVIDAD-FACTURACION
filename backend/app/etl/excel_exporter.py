"""
Utils - Excel Exporter
======================
Writes report data dicts (produced by report_service.py) into
formatted .xlsx files and returns them as bytes for Streamlit download.
"""

import io
from datetime import date

import pandas as pd
import plotly.express as px
import plotly.io as pio
from PIL import Image as PILImage
from openpyxl.drawing.image import Image as XLImage
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Theme / Style constants (improved palette for better aesthetics & contrast)
# ---------------------------------------------------------------------------

# Core palette (hex without '#')
MAIN_COLOR = "3B82F6"        # calmer bright blue (primary)
CONTRAST_COLOR = "0B2545"    # deep navy (text / header accent)
ACCENT_COLOR = "F97316"      # warm orange (call-to-action / highlights)
BACKGROUND_LIGHT = "F8FAFF"  # very light blue background for sheets
HEADER_FILL_COLOR = "E6F0FF" # soft header blue (easier on the eyes)
ROW_FILL_ODD_COLOR = "F2F7FF" # subtle alternating row fill
BORDER_COLOR = "D6E4FF"      # soft border color

# Semantic colors
SUCCESS_COLOR = "2E7D32"     # green for positive variations
DANGER_COLOR = "C62828"      # red for negative variations
NEUTRAL_COLOR = "4B5563"     # neutral gray for muted text

# Paleta para Plotly (usa las constantes ya definidas en el archivo)
PLOTLY_PALETTE = [
    f"#{MAIN_COLOR}",
    "#60A5FA",  # lighter blue
    "#A78BFA",  # purple
    f"#{ACCENT_COLOR}",
    "#34D399",  # greenish
    "#FB7185",  # pinkish
]

# Font choice
FONT_NAME = "Calibri"

# Header row of data tables
_HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
_HEADER_FONT = Font(name=FONT_NAME, bold=True, color="0B2545", size=10)

# Executive summary label column
_LABEL_FONT = Font(name=FONT_NAME, bold=True, size=10, color=CONTRAST_COLOR)
_VALUE_FONT = Font(name=FONT_NAME, size=10, color=CONTRAST_COLOR)

# Section title inside a sheet (uses accent for emphasis)
_SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color=ACCENT_COLOR)

# Cover / sheet tab accent (dark background with white text)
_ACCENT_FILL = PatternFill(fill_type="solid", fgColor=CONTRAST_COLOR)
_ACCENT_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)

# Alternating row fill and cell borders
_ROW_FILL_ODD = PatternFill(fill_type="solid", fgColor=ROW_FILL_ODD_COLOR)
_THIN_BORDER_SIDE = Side(style="thin", color=BORDER_COLOR)
_CELL_BORDER = Border(
    left=_THIN_BORDER_SIDE,
    right=_THIN_BORDER_SIDE,
    top=_THIN_BORDER_SIDE,
    bottom=_THIN_BORDER_SIDE,
)

# Variation colors
_GREEN_FONT = Font(name=FONT_NAME, bold=True, color=SUCCESS_COLOR, size=10)
_RED_FONT = Font(name=FONT_NAME, bold=True, color=DANGER_COLOR, size=10)
_NEUTRAL_FONT = Font(name=FONT_NAME, color=NEUTRAL_COLOR, size=10)



# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _apply_header_row(ws, row: int, columns: list[str]) -> None:
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _CELL_BORDER


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int = 1) -> int:
    """Write a DataFrame with header + alternating rows. Returns next empty row."""
    if df is None or df.empty:
        ws.cell(row=start_row, column=start_col, value="Sin datos disponibles").font = _NEUTRAL_FONT
        return start_row + 2

    _apply_header_row(ws, start_row, list(df.columns))

    for row_idx, row_data in enumerate(df.itertuples(index=False), start=start_row + 1):
        fill = _ROW_FILL_ODD if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=start_col):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = _VALUE_FONT
            cell.border = _CELL_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if fill:
                cell.fill = fill

    return start_row + len(df) + 2


def _auto_column_widths(ws, min_width: int = 12, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=min_width,
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max(max_len + 4, min_width), max_width
        )

def _apply_currency_format_by_header(ws, header_row: int, header_names: set[str]) -> None:
    """Apply currency format to columns identified by header name."""
    target_cols = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if isinstance(val, str) and val.strip().upper() in {h.upper() for h in header_names}:
            target_cols.append(col_idx)

    for col_idx in target_cols:
        for row_idx in range(header_row + 1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '"$"#,##0'

def _write_section_title(ws, row: int, title: str) -> None:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    cell.alignment = Alignment(vertical="center")


def _write_kpi_row(ws, row: int, label: str, value, variation_pct: float | None = None) -> None:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = _LABEL_FONT

    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.font = _VALUE_FONT
    value_cell.alignment = Alignment(horizontal="right")

    if variation_pct is not None:
        arrow = "▲" if variation_pct >= 0 else "▼"
        var_cell = ws.cell(row=row, column=3, value=f"{arrow} {variation_pct:+.1f}%")
        var_cell.font = _GREEN_FONT if variation_pct >= 0 else _RED_FONT
        var_cell.alignment = Alignment(horizontal="center")
    else:
        ws.cell(row=row, column=3, value="N/A").font = _NEUTRAL_FONT


def _write_cover_row(ws, row: int, label: str, value) -> None:
    ws.cell(row=row, column=1, value=label).font = _LABEL_FONT
    ws.cell(row=row, column=2, value=str(value)).font = _VALUE_FONT


def _format_variation(variation_block: dict) -> tuple:
    """Return (current, previous, pct) from a variation block."""
    return (
        variation_block.get("current_total", 0),
        variation_block.get("previous_total", 0),
        variation_block.get("variation_pct"),
    )


def _write_variation_rows(ws, row: int, label: str, variation_block: dict) -> int:
    current, previous, pct = _format_variation(variation_block)
    _write_kpi_row(ws, row, f"{label} (período actual)", current, variation_pct=None)
    _write_kpi_row(ws, row + 1, f"{label} (período anterior)", previous, variation_pct=None)
    _write_kpi_row(ws, row + 2, f"{label} — variación", "", variation_pct=pct)
    return row + 3


def _add_cover_info(ws, module_name: str, period_label: str) -> None:
    """Write a simple cover block at the top of the executive summary sheet."""
    ws.row_dimensions[1].height = 30
    title_cell = ws.cell(row=1, column=1, value=f"Reporte de Productividad — {module_name}")
    title_cell.font = _ACCENT_FONT
    title_cell.fill = _ACCENT_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.cell(row=2, column=1, value="Período:").font = _LABEL_FONT
    ws.cell(row=2, column=2, value=period_label).font = _VALUE_FONT
    ws.cell(row=3, column=1, value="Fecha de generación:").font = _LABEL_FONT
    ws.cell(row=3, column=2, value=str(date.today())).font = _VALUE_FONT

def _figure_to_excel_image(fig, width: int = 1200, height: int = 550, scale: int = 2) -> XLImage:
    """
    Convert Plotly figure to openpyxl image.
    Requires kaleido.
    """
    png_bytes = pio.to_image(fig, format="png", width=width, height=height, scale=scale)
    pil_img = PILImage.open(io.BytesIO(png_bytes))
    return XLImage(pil_img)


def _insert_chart(ws, fig, anchor_cell: str) -> None:
    image = _figure_to_excel_image(fig)
    ws.add_image(image, anchor_cell)

def _safe_bar(df: pd.DataFrame | None, x_col: str, y_col: str, title: str):
    if df is None or df.empty or x_col not in df.columns or y_col not in df.columns:
        return None

    n_unique = df[x_col].nunique() if x_col in df.columns else 0
    try:
        if n_unique > 1:
            fig = px.bar(
                df,
                x=x_col,
                y=y_col,
                title=title,
                text=y_col,
                color=x_col,
                color_discrete_sequence=PLOTLY_PALETTE,
            )
        else:
            fig = px.bar(
                df,
                x=x_col,
                y=y_col,
                title=title,
                text=y_col,
            )
            fig.update_traces(marker_color=f"#{MAIN_COLOR}")
    except Exception:

        fig = px.bar(df, x=x_col, y=y_col, title=title, text=y_col)

    fig.update_layout(template="plotly_white", xaxis_tickangle=-45, colorway=PLOTLY_PALETTE)
    fig.update_traces(textposition="outside")
    return fig

def _safe_line(df: pd.DataFrame | None, x_col: str, y_col: str, title: str):
    if df is None or df.empty or x_col not in df.columns or y_col not in df.columns:
        return None

    fig = px.line(df, x=x_col, y=y_col, title=title, markers=True)
    if len(fig.data) == 1:
        fig.update_traces(line=dict(color=f"#{MAIN_COLOR}"), marker=dict(color=f"#{MAIN_COLOR}"))
    fig.update_layout(template="plotly_white", xaxis_tickangle=-45, colorway=PLOTLY_PALETTE)
    return fig

def _safe_pie(df: pd.DataFrame | None, names_col: str, values_col: str, title: str):
    if df is None or df.empty or names_col not in df.columns or values_col not in df.columns:
        return None

    fig = px.pie(
        df,
        names=names_col,
        values=values_col,
        title=title,
        hole=0.35,
        color_discrete_sequence=PLOTLY_PALETTE,
    )
    fig.update_layout(template="plotly_white")
    return fig

# ---------------------------------------------------------------------------
# Per-module sheet writers
# ---------------------------------------------------------------------------

def _write_standard_executive_summary(
        ws,
        module_name: str,
        period_label: str,
        summary: dict,
) -> None:
    """
    Write executive summary sheet for billing modules.
    Layout:
        Row 1-3  : Cover info
        Row 5    : KPIs section title
        Row 6-11 : KPI rows (total, daily avg, variation blocks)
        Row 13   : Top 5 section title
        Row 14+  : Top 5 table
    """
    _add_cover_info(ws, module_name, period_label)

    _write_section_title(ws, 5, "Indicadores Clave de Productividad")

    _write_kpi_row(ws, 6, "Total procesado (período actual)", summary["total"])
    _write_kpi_row(ws, 7, "Promedio diario (período actual)", f"{summary['daily_average']:.1f}")

    row = _write_variation_rows(ws, 8, "Total procesado", summary["variation"])
    _write_variation_rows(ws, row, "Promedio diario", summary["variation_daily_avg"])

    _write_section_title(ws, 13, "Top 5 Usuarios")
    _write_dataframe(ws, summary.get("top5_by_user"), start_row=14)


def _write_standard_sheets(wb: Workbook, report: dict, module_name: str, period_label: str) -> None:
    """Add Resumen Ejecutivo, Productividad por Usuario, and Tendencia Diaria sheets."""
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    _write_standard_executive_summary(ws_summary, module_name, period_label, report["executive_summary"])
    _auto_column_widths(ws_summary)

    ws_users = wb.create_sheet("Productividad por Usuario")
    _write_section_title(ws_users, 1, "Productividad por Usuario")
    _write_dataframe(ws_users, report.get("by_user"), start_row=3)
    _auto_column_widths(ws_users)

    ws_dates = wb.create_sheet("Tendencia Diaria")
    _write_section_title(ws_dates, 1, "Tendencia Diaria")
    _write_dataframe(ws_dates, report.get("by_date"), start_row=3)
    _auto_column_widths(ws_dates)


# ---------------------------------------------------------------------------
# Public exporters
# ---------------------------------------------------------------------------

def _to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def export_billing_report(report: dict, period_label: str = "") -> bytes:
    """
    Generate billing Excel report.

    Args:
        report: Output of report_service.build_billing_report()
        period_label: Human-readable period string e.g. "01/01/2025 – 31/03/2025"

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()
    _write_standard_sheets(wb, report, "Facturación", period_label)
    # Currency format for billing tables (COUNT = valor facturado)
    ws_users = wb["Productividad por Usuario"]
    ws_dates = wb["Tendencia Diaria"]
    ws_summary = wb["Resumen Ejecutivo"]

    _apply_currency_format_by_header(ws_users, header_row=3, header_names={"COUNT"})
    _apply_currency_format_by_header(ws_dates, header_row=3, header_names={"COUNT"})
    _apply_currency_format_by_header(ws_summary, header_row=14, header_names={"COUNT"})

    ws_records = wb.create_sheet("Tendencia Registros")
    _write_section_title(ws_records, 1, "Cantidad de Registros")
    _write_dataframe(ws_records, report.get("by_date_records"), start_row=3)
    _auto_column_widths(ws_records)


    if isinstance(ws_summary["B6"].value, (int, float)):
        ws_summary["B6"].number_format = '"$"#,##0'

    _add_standard_charts_sheet(wb, report, "Facturación")
    return _to_bytes(wb)


def export_billing_report_cached(report: dict, period_label: str = "") -> bytes:
    """Cached wrapper for billing Excel export."""
    return export_billing_report(report, period_label=period_label)


def export_legalizations_report(report: dict, period_label: str = "") -> bytes:
    """
    Generate legalizations Excel report with PPL and Agreements sheets.

    Args:
        report: Output of report_service.build_legalizations_report()
        period_label: Human-readable period string.

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()

    # --- Resumen Ejecutivo ---
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    _add_cover_info(ws_summary, "Legalizaciones", period_label)

    summary = report["executive_summary"]
    _write_section_title(ws_summary, 5, "Indicadores Globales")
    _write_kpi_row(ws_summary, 6, "Total global (PPL + Convenios)", summary["total"])
    _write_kpi_row(ws_summary, 7, "Total PPL", summary["ppl_total"])
    _write_kpi_row(ws_summary, 8, "Total Convenios", summary["agreements_total"])
    _write_kpi_row(ws_summary, 9, "Promedio diario global", f"{summary['daily_average']:.1f}")

    row = _write_variation_rows(ws_summary, 10, "Total global", summary["variation"])
    _write_variation_rows(ws_summary, row, "Promedio diario", summary["variation_daily_avg"])

    _write_section_title(ws_summary, row + 3, "Top 5 Usuarios — PPL")
    next_row = _write_dataframe(ws_summary, report["ppl"].get("top5_by_user"), start_row=row + 4)
    _write_section_title(ws_summary, next_row, "Top 5 Usuarios — Convenios")
    _write_dataframe(ws_summary, report["agreements"].get("top5_by_user"), start_row=next_row + 1)
    _auto_column_widths(ws_summary)

    # --- PPL sheets ---
    ws_ppl_users = wb.create_sheet("PPL - Por Usuario")
    _write_section_title(ws_ppl_users, 1, "Legalizaciones PPL — Por Usuario")
    _write_dataframe(ws_ppl_users, report["ppl"].get("by_user"), start_row=3)
    _auto_column_widths(ws_ppl_users)

    ws_ppl_dates = wb.create_sheet("PPL - Tendencia")
    _write_section_title(ws_ppl_dates, 1, "Legalizaciones PPL — Tendencia Diaria")
    _write_dataframe(ws_ppl_dates, report["ppl"].get("by_date"), start_row=3)
    _auto_column_widths(ws_ppl_dates)

    # --- Agreements sheets ---
    ws_conv_users = wb.create_sheet("Convenios - Por Usuario")
    _write_section_title(ws_conv_users, 1, "Legalizaciones Convenios — Por Usuario")
    _write_dataframe(ws_conv_users, report["agreements"].get("by_user"), start_row=3)
    _auto_column_widths(ws_conv_users)

    ws_conv_dates = wb.create_sheet("Convenios - Tendencia")
    _write_section_title(ws_conv_dates, 1, "Legalizaciones Convenios — Tendencia Diaria")
    _write_dataframe(ws_conv_dates, report["agreements"].get("by_date"), start_row=3)
    _auto_column_widths(ws_conv_dates)

    _add_legalizations_charts_sheet(wb, report)
    return _to_bytes(wb)


def export_legalizations_report_cached(report: dict, period_label: str = "") -> bytes:
    """Cached wrapper for legalizations Excel export."""
    return export_legalizations_report(report, period_label=period_label)


def export_processes_report(report: dict, period_label: str = "") -> bytes:
    """
    Generate administrative processes Excel report.

    Args:
        report: Output of report_service.build_processes_report()
        period_label: Human-readable period string.

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()

    # --- Resumen Ejecutivo ---
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    _add_cover_info(ws_summary, "Procesos Administrativos", period_label)

    summary = report["executive_summary"]
    _write_section_title(ws_summary, 5, "Indicadores Clave")
    _write_kpi_row(ws_summary, 6, "Total registros", summary["total_records"])
    _write_kpi_row(ws_summary, 7, "Total cantidad", f"{summary['total_quantity']:,.0f}")
    _write_kpi_row(ws_summary, 8, "Personas activas", summary["unique_people"])
    _write_kpi_row(ws_summary, 9, "Tipos de proceso activos", summary["unique_processes"])

    row = _write_variation_rows(ws_summary, 10, "Total registros", summary["variation_records"])
    _write_variation_rows(ws_summary, row, "Total cantidad", summary["variation_quantity"])
    _auto_column_widths(ws_summary)

    # --- Por Persona ---
    ws_person = wb.create_sheet("Por Persona")
    _write_section_title(ws_person, 1, "Productividad por Persona")
    _write_dataframe(ws_person, report.get("by_person"), start_row=3)
    _auto_column_widths(ws_person)

    # --- Por Proceso ---
    ws_process = wb.create_sheet("Por Proceso")
    _write_section_title(ws_process, 1, "Productividad por Tipo de Proceso")
    _write_dataframe(ws_process, report.get("by_process"), start_row=3)
    _auto_column_widths(ws_process)

    # --- Tendencia Diaria ---
    ws_trend = wb.create_sheet("Tendencia Diaria")
    _write_section_title(ws_trend, 1, "Tendencia Diaria")
    time_trend = report.get("chart_datasets", {}).get("time_trend")
    _write_dataframe(ws_trend, time_trend, start_row=3)
    _auto_column_widths(ws_trend)

    _add_processes_charts_sheet(wb, report)
    return _to_bytes(wb)


def export_processes_report_cached(report: dict, period_label: str = "") -> bytes:
    """Cached wrapper for administrative processes Excel export."""
    return export_processes_report(report, period_label=period_label)


# ---------------------------------------------------------------------------
# Graphic exporters
# -------

def _add_standard_charts_sheet(wb: Workbook, report: dict, module_name: str) -> None:
    ws = wb.create_sheet("Graficos")

    by_user = report.get("by_user")                  # valor facturado por usuario
    by_date = report.get("by_date")                  # valor facturado por fecha
    by_user_records = report.get("by_user_records")  # registros por usuario

    if by_user is not None and not by_user.empty:
        user_col = by_user.columns[0]
        fig_user = _safe_bar(by_user, user_col, "COUNT", f"{module_name} - Valor por Usuario")
        if fig_user:
            _insert_chart(ws, fig_user, "A1")

    if by_user_records is not None and not by_user_records.empty:
        user_col_records = by_user_records.columns[0]
        fig_records_user = _safe_bar(
            by_user_records,
            user_col_records,
            "REGISTROS",
            f"{module_name} - Registros por Usuario",
        )
        if fig_records_user:
            _insert_chart(ws, fig_records_user, "M1")

    fig_value_date = _safe_line(
        by_date,
        "DATE",
        "COUNT",
        f"{module_name} - Valor por Fecha",
    )
    if fig_value_date:
        _insert_chart(ws, fig_value_date, "A30")


def _add_legalizations_charts_sheet(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Graficos")

    ppl_user = report.get("ppl", {}).get("by_user")
    ppl_date = report.get("ppl", {}).get("by_date")
    agr_user = report.get("agreements", {}).get("by_user")
    agr_date = report.get("agreements", {}).get("by_date")

    if ppl_user is not None and not ppl_user.empty:
        ppl_user_col = ppl_user.columns[0]
        fig = _safe_bar(ppl_user, ppl_user_col, "REGISTROS", "PPL por Usuario")
        if fig:
            _insert_chart(ws, fig, "A1")

    if ppl_date is not None and not ppl_date.empty:
        fig = _safe_line(ppl_date, "DATE", "REGISTROS", "Tendencia PPL")
        if fig:
            _insert_chart(ws, fig, "A30")

    if agr_user is not None and not agr_user.empty:
        agr_user_col = agr_user.columns[0]
        fig = _safe_bar(agr_user, agr_user_col, "REGISTROS", "Convenios por Usuario")
        if fig:
            _insert_chart(ws, fig, "M1")

    if agr_date is not None and not agr_date.empty:
        fig = _safe_line(agr_date, "DATE", "REGISTROS", "Tendencia Convenios")
        if fig:
            _insert_chart(ws, fig, "M30")

def _add_processes_charts_sheet(wb: Workbook, report: dict) -> None:
    ws = wb.create_sheet("Graficos")
    charts = report.get("chart_datasets", {})

    bar_by_person = charts.get("bar_by_person")
    pie_distribution = charts.get("pie_distribution")
    pie_mode = charts.get("pie_mode")
    time_trend = charts.get("time_trend")

    fig_bar = _safe_bar(bar_by_person, "NOMBRE", "CANTIDAD", "Cantidad por Persona")
    if fig_bar:
        _insert_chart(ws, fig_bar, "A1")

    if pie_mode == "person":
        fig_pie = _safe_pie(pie_distribution, "NOMBRE", "CANTIDAD", "Distribucion por Persona")
    else:
        fig_pie = _safe_pie(pie_distribution, "PROCESO", "CANTIDAD", "Cantidad por Proceso")
    if fig_pie:
        _insert_chart(ws, fig_pie, "M1")

    fig_line = _safe_line(time_trend, "FECHA", "CANTIDAD", "Tendencia Temporal")
    if fig_line:
        _insert_chart(ws, fig_line, "A30")


# ---------------------------------------------------------------------------
# RIPS exporter
# ---------------------------------------------------------------------------

def export_rips_report(report: dict, period_label: str = "") -> bytes:
    """
    Generate RIPS Excel report.

    Args:
        report: Output of report_service.build_rips_report()
        period_label: Human-readable period string.

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()

    # --- Resumen Ejecutivo ---
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    _add_cover_info(ws_summary, "RIPS", period_label)

    summary = report["executive_summary"]
    _write_section_title(ws_summary, 5, "Indicadores Clave")
    _write_kpi_row(ws_summary, 6, "Total RIPS completados", summary["total"])
    _write_kpi_row(ws_summary, 7, "Promedio diario", f"{summary['daily_average']:.1f}")

    row = _write_variation_rows(ws_summary, 8, "Total RIPS", summary["variation"])
    _write_variation_rows(ws_summary, row, "Promedio diario", summary["variation_daily_avg"])

    _write_section_title(ws_summary, row + 3, "Top 5 Usuarios")
    _write_dataframe(ws_summary, summary.get("top5_by_user"), start_row=row + 4)
    _auto_column_widths(ws_summary)

    # --- Por Usuario ---
    ws_users = wb.create_sheet("Por Usuario")
    _write_section_title(ws_users, 1, "Productividad por Usuario")
    _write_dataframe(ws_users, report.get("by_user"), start_row=3)
    _auto_column_widths(ws_users)

    # --- Tendencia Diaria ---
    ws_dates = wb.create_sheet("Tendencia Diaria")
    _write_section_title(ws_dates, 1, "Tendencia Diaria")
    _write_dataframe(ws_dates, report.get("by_date"), start_row=3)
    _auto_column_widths(ws_dates)

    # --- Gráficos ---
    ws_charts = wb.create_sheet("Graficos")
    by_user = report.get("by_user")
    by_date = report.get("by_date")

    if by_user is not None and not by_user.empty:
        user_col = by_user.columns[0]
        fig_user = _safe_bar(by_user, user_col, "REGISTROS", "RIPS por Usuario")
        if fig_user:
            _insert_chart(ws_charts, fig_user, "A1")

    fig_date = _safe_line(by_date, "DATE", "REGISTROS", "Tendencia RIPS")
    if fig_date:
        _insert_chart(ws_charts, fig_date, "A30")

    return _to_bytes(wb)


def export_rips_report_cached(report: dict, period_label: str = "") -> bytes:
    """Cached wrapper for RIPS Excel export."""
    return export_rips_report(report, period_label=period_label)


# ---------------------------------------------------------------------------
# Radicación exporter
# ---------------------------------------------------------------------------

def export_radicacion_report(report: dict, period_label: str = "") -> bytes:
    """
    Generate Radicación Excel report.

    Args:
        report: Output of report_service.build_radicacion_report()
        period_label: Human-readable period string.

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()

    # --- Resumen Ejecutivo ---
    ws_summary = wb.active
    ws_summary.title = "Resumen Ejecutivo"
    _add_cover_info(ws_summary, "Radicación", period_label)

    summary = report["executive_summary"]
    _write_section_title(ws_summary, 5, "Indicadores Clave")
    _write_kpi_row(ws_summary, 6, "Total facturas", summary["total"])
    _write_kpi_row(ws_summary, 7, "Radicadas", summary["radicadas"])
    _write_kpi_row(ws_summary, 8, "Vencidas", summary["vencidas"])
    _write_kpi_row(ws_summary, 9, "% Radicado", f"{summary['pct_radicado']:.1f}%")
    _write_kpi_row(ws_summary, 10, "% Vencidas", f"{summary['pct_vencidas']:.1f}%")

    row = _write_variation_rows(ws_summary, 11, "Total facturas", summary["variation"])
    _write_variation_rows(ws_summary, row, "Facturas vencidas", summary["variation_vencidas"])
    _auto_column_widths(ws_summary)

    # --- Facturas Vencidas ---
    ws_vencidas = wb.create_sheet("Facturas Vencidas")
    _write_section_title(ws_vencidas, 1, "Facturas Vencidas (sin radicar > 2 días)")
    _write_dataframe(ws_vencidas, report.get("vencidas_df"), start_row=3)
    _auto_column_widths(ws_vencidas)

    # --- Por Usuario ---
    ws_users = wb.create_sheet("Por Usuario")
    _write_section_title(ws_users, 1, "Radicación por Usuario")
    _write_dataframe(ws_users, report.get("by_user"), start_row=3)
    _auto_column_widths(ws_users)

    # --- Gráficos ---
    ws_charts = wb.create_sheet("Graficos")
    by_user = report.get("by_user")

    if by_user is not None and not by_user.empty:
        fig_total = _safe_bar(by_user, "USUARIO", "TOTAL", "Total Facturas por Usuario")
        if fig_total:
            _insert_chart(ws_charts, fig_total, "A1")

        fig_vencidas = _safe_bar(by_user, "USUARIO", "VENCIDAS", "Facturas Vencidas por Usuario")
        if fig_vencidas:
            _insert_chart(ws_charts, fig_vencidas, "M1")

        fig_pie = _safe_pie(by_user, "USUARIO", "VENCIDAS", "Distribución de Vencidas")
        if fig_pie:
            _insert_chart(ws_charts, fig_pie, "A30")

    return _to_bytes(wb)


def export_radicacion_report_cached(report: dict, period_label: str = "") -> bytes:
    """Cached wrapper for radicación Excel export."""
    return export_radicacion_report(report, period_label=period_label)


# ---------------------------------------------------------------------------
# General report exporter (combines all modules)
# ---------------------------------------------------------------------------

def export_general_report(reports: dict, period_label: str = "") -> bytes:
    """
    Generate a combined general Excel report with all modules.

    Args:
        reports: Output of report_service.build_general_report()
                 Keys: billing, legalizations, rips, radicacion, processes
        period_label: Human-readable period string.

    Returns:
        bytes of the .xlsx file.
    """
    wb = Workbook()

    # --- Portada ---
    ws_cover = wb.active
    ws_cover.title = "Portada"
    ws_cover.row_dimensions[1].height = 40
    title_cell = ws_cover.cell(row=1, column=1, value="Informe General de Productividad")
    title_cell.font = _ACCENT_FONT
    title_cell.fill = _ACCENT_FILL
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    ws_cover.cell(row=3, column=1, value="Período:").font = _LABEL_FONT
    ws_cover.cell(row=3, column=2, value=period_label).font = _VALUE_FONT
    ws_cover.cell(row=4, column=1, value="Fecha de generación:").font = _LABEL_FONT
    ws_cover.cell(row=4, column=2, value=str(date.today())).font = _VALUE_FONT

    ws_cover.cell(row=6, column=1, value="Módulos incluidos:").font = _SECTION_FONT
    modules_list = []
    if reports.get("billing"):
        modules_list.append("Facturación Electrónica")
    if reports.get("legalizations"):
        modules_list.append("Legalizaciones")
    if reports.get("rips"):
        modules_list.append("RIPS")
    if reports.get("radicacion"):
        modules_list.append("Radicación")
    if reports.get("processes"):
        modules_list.append("Procesos Administrativos")

    for i, mod in enumerate(modules_list, start=7):
        ws_cover.cell(row=i, column=1, value=f"• {mod}").font = _VALUE_FONT

    _auto_column_widths(ws_cover)

    # --- Resumen Ejecutivo Global ---
    ws_global = wb.create_sheet("Resumen Global")
    _write_section_title(ws_global, 1, "Resumen Ejecutivo Global")

    row = 3
    total_all = 0

    billing_rep = reports.get("billing")
    if billing_rep:
        es = billing_rep.get("executive_summary", {})
        _write_kpi_row(ws_global, row, "Facturación — Total", es.get("total", 0))
        _write_kpi_row(ws_global, row + 1, "Facturación — Valor", f"${es.get('total', 0):,.0f}")
        row += 3
        total_all += es.get("total", 0)

    leg_rep = reports.get("legalizations")
    if leg_rep:
        es = leg_rep.get("executive_summary", {})
        _write_kpi_row(ws_global, row, "Legalizaciones — Total", es.get("total", 0))
        _write_kpi_row(ws_global, row + 1, "Legalizaciones — PPL", es.get("ppl_total", 0))
        _write_kpi_row(ws_global, row + 2, "Legalizaciones — Convenios", es.get("agreements_total", 0))
        row += 4
        total_all += es.get("total", 0)

    rips_rep = reports.get("rips")
    if rips_rep:
        es = rips_rep.get("executive_summary", {})
        _write_kpi_row(ws_global, row, "RIPS — Total", es.get("total", 0))
        row += 2
        total_all += es.get("total", 0)

    rad_rep = reports.get("radicacion")
    if rad_rep:
        es = rad_rep.get("executive_summary", {})
        _write_kpi_row(ws_global, row, "Radicación — Total", es.get("total", 0))
        _write_kpi_row(ws_global, row + 1, "Radicación — Vencidas", es.get("vencidas", 0))
        _write_kpi_row(ws_global, row + 2, "Radicación — % Radicado", f"{es.get('pct_radicado', 0):.1f}%")
        row += 4
        total_all += es.get("total", 0)

    proc_rep = reports.get("processes")
    if proc_rep:
        es = proc_rep.get("executive_summary", {})
        _write_kpi_row(ws_global, row, "Procesos — Total registros", es.get("total_records", 0))
        _write_kpi_row(ws_global, row + 1, "Procesos — Total cantidad", f"{es.get('total_quantity', 0):,.0f}")
        row += 3
        total_all += es.get("total_records", 0)

    _write_kpi_row(ws_global, row, "TOTAL GENERAL", total_all)
    _auto_column_widths(ws_global)

    # --- Facturación sheets ---
    if billing_rep:
        _add_billing_sheets_to_general(wb, billing_rep, period_label)
        _add_billing_charts_to_general(wb, billing_rep)

    # --- Legalizaciones sheets ---
    if leg_rep:
        _add_legalizations_sheets(wb, leg_rep, period_label)
        _add_legalizations_charts_to_general(wb, leg_rep)

    # --- RIPS sheets ---
    if rips_rep:
        _add_rips_sheets(wb, rips_rep, period_label)
        _add_rips_charts_to_general(wb, rips_rep)

    # --- Radicación sheets ---
    if rad_rep:
        _add_radicacion_sheets(wb, rad_rep, period_label)
        _add_radicacion_charts_to_general(wb, rad_rep)

    # --- Procesos sheets ---
    if proc_rep:
        _add_processes_sheets(wb, proc_rep, period_label)
        _add_processes_charts_to_general(wb, proc_rep)

    return _to_bytes(wb)


def export_general_report_cached(reports: dict, period_label: str = "") -> bytes:
    """Cached wrapper for general Excel export."""
    return export_general_report(reports, period_label=period_label)


# ---------------------------------------------------------------------------
# Helper: add legalizations sheets to existing workbook
# ---------------------------------------------------------------------------

def _add_legalizations_sheets(wb: Workbook, report: dict, period_label: str) -> None:
    """Unified legalizations sheet with all data."""
    summary = report["executive_summary"]

    ws = wb.create_sheet("Legalizaciones")
    _add_cover_info(ws, "Legalizaciones", period_label)

    # --- KPIs Section ---
    _write_section_title(ws, 5, "Indicadores Globales")
    _write_kpi_row(ws, 6, "Total global (PPL + Convenios)", summary["total"])
    _write_kpi_row(ws, 7, "Total PPL", summary["ppl_total"])
    _write_kpi_row(ws, 8, "Total Convenios", summary["agreements_total"])
    row = _write_variation_rows(ws, 9, "Total global", summary["variation"])

    # --- PPL Section ---
    _write_section_title(ws, row + 2, "PPL — Top 5 Usuarios")
    row = _write_dataframe(ws, report["ppl"].get("top5_by_user"), start_row=row + 3)

    _write_section_title(ws, row, "PPL — Productividad por Usuario")
    row = _write_dataframe(ws, report["ppl"].get("by_user"), start_row=row + 1)

    _write_section_title(ws, row, "PPL — Tendencia Diaria")
    row = _write_dataframe(ws, report["ppl"].get("by_date"), start_row=row + 1)

    # --- Convenios Section ---
    _write_section_title(ws, row + 2, "Convenios — Top 5 Usuarios")
    row = _write_dataframe(ws, report["agreements"].get("top5_by_user"), start_row=row + 3)

    _write_section_title(ws, row, "Convenios — Productividad por Usuario")
    row = _write_dataframe(ws, report["agreements"].get("by_user"), start_row=row + 1)

    _write_section_title(ws, row, "Convenios — Tendencia Diaria")
    _write_dataframe(ws, report["agreements"].get("by_date"), start_row=row + 1)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Helper: add RIPS sheets to existing workbook
# ---------------------------------------------------------------------------

def _add_rips_sheets(wb: Workbook, report: dict, period_label: str) -> None:
    summary = report["executive_summary"]

    ws = wb.create_sheet("RIPS")
    _add_cover_info(ws, "RIPS", period_label)

    _write_section_title(ws, 5, "Indicadores Clave")
    _write_kpi_row(ws, 6, "Total RIPS", summary["total"])
    _write_kpi_row(ws, 7, "Promedio diario", f"{summary['daily_average']:.1f}")
    row = _write_variation_rows(ws, 8, "Total RIPS", summary["variation"])

    _write_section_title(ws, row + 2, "Top 5 Usuarios")
    row = _write_dataframe(ws, summary.get("top5_by_user"), start_row=row + 3)

    _write_section_title(ws, row, "Productividad por Usuario")
    row = _write_dataframe(ws, report.get("by_user"), start_row=row + 1)

    _write_section_title(ws, row, "Tendencia Diaria")
    _write_dataframe(ws, report.get("by_date"), start_row=row + 1)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Helper: add Radicación sheets to existing workbook
# ---------------------------------------------------------------------------

def _add_radicacion_sheets(wb: Workbook, report: dict, period_label: str) -> None:
    summary = report["executive_summary"]

    ws = wb.create_sheet("Radicación")
    _add_cover_info(ws, "Radicación", period_label)

    _write_section_title(ws, 5, "Indicadores Clave")
    _write_kpi_row(ws, 6, "Total facturas", summary["total"])
    _write_kpi_row(ws, 7, "Radicadas", summary["radicadas"])
    _write_kpi_row(ws, 8, "Vencidas", summary["vencidas"])
    _write_kpi_row(ws, 9, "% Radicado", f"{summary['pct_radicado']:.1f}%")
    row = _write_variation_rows(ws, 10, "Total facturas", summary["variation"])
    _write_variation_rows(ws, row, "Facturas vencidas", summary["variation_vencidas"])

    _write_section_title(ws, row + 2, "Facturas Vencidas")
    row = _write_dataframe(ws, report.get("vencidas_df"), start_row=row + 3)

    _write_section_title(ws, row, "Radicación por Usuario")
    _write_dataframe(ws, report.get("by_user"), start_row=row + 1)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Helper: add Processes sheets to existing workbook
# ---------------------------------------------------------------------------

def _add_processes_sheets(wb: Workbook, report: dict, period_label: str) -> None:
    summary = report["executive_summary"]

    ws = wb.create_sheet("Procesos Administrativos")
    _add_cover_info(ws, "Procesos Administrativos", period_label)

    _write_section_title(ws, 5, "Indicadores Clave")
    _write_kpi_row(ws, 6, "Total registros", summary["total_records"])
    _write_kpi_row(ws, 7, "Total cantidad", f"{summary['total_quantity']:,.0f}")
    _write_kpi_row(ws, 8, "Personas activas", summary["unique_people"])
    row = _write_variation_rows(ws, 9, "Total registros", summary["variation_records"])
    _write_variation_rows(ws, row, "Total cantidad", summary["variation_quantity"])

    _write_section_title(ws, row + 2, "Por Persona")
    row = _write_dataframe(ws, report.get("by_person"), start_row=row + 3)

    _write_section_title(ws, row, "Por Proceso")
    row = _write_dataframe(ws, report.get("by_process"), start_row=row + 1)

    _write_section_title(ws, row, "Tendencia Diaria")
    time_trend = report.get("chart_datasets", {}).get("time_trend")
    _write_dataframe(ws, time_trend, start_row=row + 1)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Helper: add billing sheets to existing workbook
# ---------------------------------------------------------------------------

def _add_billing_sheets_to_general(wb: Workbook, report: dict, period_label: str) -> None:
    summary = report["executive_summary"]

    ws = wb.create_sheet("Facturación Electrónica")
    _add_cover_info(ws, "Facturación Electrónica", period_label)

    _write_section_title(ws, 5, "Indicadores Clave de Productividad")
    _write_kpi_row(ws, 6, "Total procesado", summary["total"])
    _write_kpi_row(ws, 7, "Promedio diario", f"{summary['daily_average']:.1f}")
    row = _write_variation_rows(ws, 8, "Total procesado", summary["variation"])
    _write_variation_rows(ws, row, "Promedio diario", summary["variation_daily_avg"])

    _write_section_title(ws, row + 2, "Top 5 Usuarios")
    row = _write_dataframe(ws, summary.get("top5_by_user"), start_row=row + 3)

    _write_section_title(ws, row, "Productividad por Usuario")
    row = _write_dataframe(ws, report.get("by_user"), start_row=row + 1)

    _write_section_title(ws, row, "Tendencia Diaria (Valor)")
    row = _write_dataframe(ws, report.get("by_date"), start_row=row + 1)

    _write_section_title(ws, row, "Tendencia Diaria (Registros)")
    _write_dataframe(ws, report.get("by_date_records"), start_row=row + 1)

    _auto_column_widths(ws)


# ---------------------------------------------------------------------------
# Chart helpers for general report
# ---------------------------------------------------------------------------

def _add_billing_charts_to_general(wb: Workbook, report: dict) -> None:
    """Add billing charts to general report."""
    ws = wb.create_sheet("Facturación - Gráficos")

    by_user = report.get("by_user")
    by_date = report.get("by_date")
    by_user_records = report.get("by_user_records")

    if by_user is not None and not by_user.empty:
        user_col = by_user.columns[0]
        fig_user = _safe_bar(by_user, user_col, "COUNT", "Facturación - Valor por Usuario")
        if fig_user:
            _insert_chart(ws, fig_user, "A1")

    if by_user_records is not None and not by_user_records.empty:
        user_col_records = by_user_records.columns[0]
        fig_records = _safe_bar(by_user_records, user_col_records, "REGISTROS", "Facturación - Registros por Usuario")
        if fig_records:
            _insert_chart(ws, fig_records, "M1")

    if by_date is not None and not by_date.empty:
        fig_date = _safe_line(by_date, "DATE", "COUNT", "Facturación - Tendencia de Valor")
        if fig_date:
            _insert_chart(ws, fig_date, "A30")


def _add_legalizations_charts_to_general(wb: Workbook, report: dict) -> None:
    """Add legalizations charts to general report."""
    ws = wb.create_sheet("Legalizaciones - Gráficos")

    ppl_user = report.get("ppl", {}).get("by_user")
    ppl_date = report.get("ppl", {}).get("by_date")
    agr_user = report.get("agreements", {}).get("by_user")
    agr_date = report.get("agreements", {}).get("by_date")

    if ppl_user is not None and not ppl_user.empty:
        ppl_user_col = ppl_user.columns[0]
        fig = _safe_bar(ppl_user, ppl_user_col, "REGISTROS", "PPL por Usuario")
        if fig:
            _insert_chart(ws, fig, "A1")

    if ppl_date is not None and not ppl_date.empty:
        fig = _safe_line(ppl_date, "DATE", "REGISTROS", "Tendencia PPL")
        if fig:
            _insert_chart(ws, fig, "A30")

    if agr_user is not None and not agr_user.empty:
        agr_user_col = agr_user.columns[0]
        fig = _safe_bar(agr_user, agr_user_col, "REGISTROS", "Convenios por Usuario")
        if fig:
            _insert_chart(ws, fig, "M1")

    if agr_date is not None and not agr_date.empty:
        fig = _safe_line(agr_date, "DATE", "REGISTROS", "Tendencia Convenios")
        if fig:
            _insert_chart(ws, fig, "M30")


def _add_rips_charts_to_general(wb: Workbook, report: dict) -> None:
    """Add RIPS charts to general report."""
    ws = wb.create_sheet("RIPS - Gráficos")

    by_user = report.get("by_user")
    by_date = report.get("by_date")

    if by_user is not None and not by_user.empty:
        user_col = by_user.columns[0]
        fig_user = _safe_bar(by_user, user_col, "REGISTROS", "RIPS por Usuario")
        if fig_user:
            _insert_chart(ws, fig_user, "A1")

    if by_date is not None and not by_date.empty:
        fig_date = _safe_line(by_date, "DATE", "REGISTROS", "Tendencia RIPS")
        if fig_date:
            _insert_chart(ws, fig_date, "A30")


def _add_radicacion_charts_to_general(wb: Workbook, report: dict) -> None:
    """Add Radicación charts to general report."""
    ws = wb.create_sheet("Radicación - Gráficos")

    by_user = report.get("by_user")

    if by_user is not None and not by_user.empty:
        fig_total = _safe_bar(by_user, "USUARIO", "TOTAL", "Total Facturas por Usuario")
        if fig_total:
            _insert_chart(ws, fig_total, "A1")

        fig_vencidas = _safe_bar(by_user, "USUARIO", "VENCIDAS", "Facturas Vencidas por Usuario")
        if fig_vencidas:
            _insert_chart(ws, fig_vencidas, "M1")

        fig_pie = _safe_pie(by_user, "USUARIO", "VENCIDAS", "Distribución de Vencidas")
        if fig_pie:
            _insert_chart(ws, fig_pie, "A30")


def _add_processes_charts_to_general(wb: Workbook, report: dict) -> None:
    """Add Processes charts to general report."""
    ws = wb.create_sheet("Procesos - Gráficos")

    charts = report.get("chart_datasets", {})
    bar_by_person = charts.get("bar_by_person")
    pie_distribution = charts.get("pie_distribution")
    pie_mode = charts.get("pie_mode")
    time_trend = charts.get("time_trend")

    if bar_by_person is not None and not bar_by_person.empty:
        fig_bar = _safe_bar(bar_by_person, "NOMBRE", "CANTIDAD", "Cantidad por Persona")
        if fig_bar:
            _insert_chart(ws, fig_bar, "A1")

    if pie_distribution is not None and not pie_distribution.empty:
        if pie_mode == "person":
            fig_pie = _safe_pie(pie_distribution, "NOMBRE", "CANTIDAD", "Distribución por Persona")
        else:
            fig_pie = _safe_pie(pie_distribution, "PROCESO", "CANTIDAD", "Distribución por Proceso")
        if fig_pie:
            _insert_chart(ws, fig_pie, "M1")

    if time_trend is not None and not time_trend.empty:
        fig_line = _safe_line(time_trend, "FECHA", "CANTIDAD", "Tendencia Temporal")
        if fig_line:
            _insert_chart(ws, fig_line, "A30")
